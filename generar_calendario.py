#!/usr/bin/env python3
"""
Generador de Calendario de Disponibilidad
=========================================
Crea un .xlsx con todas las hojas (Calendario, Resumen, Julio, Agosto, Cuadricula, Instrucciones)
mas el .gs de Apps Script.

USO:
    python3 generar_calendario.py

Personaliza las variables del bloque CONFIGURACION abajo.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from datetime import date, timedelta

# ============================================================
# CONFIGURACION  —  cambia esto a tu gusto
# ============================================================
YEAR = 2026
MONTHS = [7, 8]  # numeros de mes (1-12)
MONTH_NAMES = {7: "JULIO", 8: "AGOSTO"}  # nombre para cada mes

PEOPLE = [
    "ELIAS",
    "MERINO",
    "PONSA",
    "FERRAN",
    "AUGUST",
    "JOAN",
    "GRAU",
    "POL",
]

CODES = ["TM", "TT", "TN", "FV", "FN", "OC", "RE", "OT", "CL"]

# Significado de cada codigo (para la leyenda)
CODE_MEANINGS = {
    "TM": "Trabajar manana (~08-15)",
    "TT": "Trabajar tarde (~15-21)",
    "TN": "Trabajar noche (~21-08)",
    "FV": "Fuera (puedo volver)",
    "FN": "Fuera (NO volver)",
    "OC": "Ocupado",
    "RE": "Recuperaciones",
    "OT": "Otros",
    "CL": "Clases",
}

# Colores para cada codigo: (fondo_hex, texto_hex, negrita)
CODE_COLORS = {
    "TM": ("B6D7A8", "1B2631", False),  # verde claro
    "TT": ("6AA84F", "FFFFFF", False),  # verde medio
    "TN": ("38761D", "FFFFFF", True),  # verde oscuro
    "FV": ("F6B26B", "1B2631", False),  # naranja
    "FN": ("E06666", "FFFFFF", True),  # rojo
    "OC": ("717D7E", "FFFFFF", True),  # gris
    "RE": ("8E44AD", "FFFFFF", True),  # purpura
    "OT": ("17A589", "FFFFFF", True),  # teal
    "CL": ("E67E22", "FFFFFF", True),  # naranja oscuro / clases
}

# Emoji para cuadricula (se renderizan en color)
EMOJI = {
    "free": "🟢",
    "TM": "🔵",
    "TT": "🟠",
    "TN": "🟣",
    "FV": "🟡",
    "CL": "🟤",
}

# ============================================================
# NO TOCAR NADA DE AQUI PARA ABAJO
# ============================================================

# Codigos que NO aparecen en la cuadricula (100% no disponibles)
HIDDEN_IN_GRID = {"FN", "OC", "RE", "OT"}

# Codigos "disponibles" que aparecen en la cuadricula
SHOWN_IN_GRID = {"TM", "TT", "TN", "FV", "CL"}


def generate(year, months, month_names, people, group_name="grupo"):
    group_slug = group_name.replace(" ", "_")
    YEAR = year
    MONTHS = months
    MONTH_NAMES = month_names
    PEOPLE = people

    wb = openpyxl.Workbook()
    DARK = "1B2631"
    WHITE = "FFFFFF"
    WEEKEND_BG = "F2F3F4"
    ALT_ROW = "F8F9FA"

    thin = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )
    center = Alignment(horizontal="center", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    hdr_font = Font(bold=True, color=WHITE, size=11)

    def style_header(ws, headers, row=1, fill=None, font=None):
        f = fill or hdr_fill
        fn = font or hdr_font
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = fn
            c.fill = f
            c.alignment = center_wrap
            c.border = thin
        ws.row_dimensions[row].height = 32

    def set_col_widths(ws, widths):
        for c, w in widths:
            ws.column_dimensions[c].width = w

    # Dias de la semana
    day_n = {0: "Lun", 1: "Mar", 2: "Mie", 3: "Jue", 4: "Vie", 5: "Sab", 6: "Dom"}

    # Calcular total de dias y fechas
    total_days = 0
    month_dates = {}
    for m in MONTHS:
        if m == 12:
            end_d = date(YEAR, 12, 31)
        else:
            end_d = date(YEAR, m + 1, 1) - timedelta(days=1)
        month_dates[m] = end_d.day
        total_days += end_d.day

    # Paleta de colores para los meses
    month_palette = [
        ("2980B9", "AED6F1", "D6EAF8", "1A5276"),  # azul
        ("1E8449", "A9DFBF", "D5F5E3", "0E6655"),  # verde
        ("8E44AD", "D2B4DE", "E8DAEF", "6C3483"),  # purpura
        ("E67E22", "F0B27A", "FDEBD0", "CA6F1E"),  # naranja
        ("C0392B", "F1948A", "FDEDEC", "A93226"),  # rojo
        ("2E86C1", "85C1E9", "D6EAF8", "1B4F72"),  # azul oscuro
    ]

    # ============================================================
    # SHEET 1: CALENDARIO
    # ============================================================
    ws = wb.active
    ws.title = "Calendario"
    ws.sheet_properties.tabColor = month_palette[0][0]

    cal_widths = [("A", 14), ("B", 7)] + [
        (get_column_letter(i), 10) for i in range(3, 3 + len(PEOPLE))
    ]
    set_col_widths(ws, cal_widths)

    style_header(ws, ["Fecha", "Dia"] + PEOPLE)

    curr = date(YEAR, MONTHS[0], 1)
    end = date(YEAR, MONTHS[-1], month_dates[MONTHS[-1]])
    row = 2
    cal_map = []
    month_index = 0

    while curr <= end:
        if curr.day == 1 and month_index < len(MONTHS):
            mn = MONTH_NAMES.get(curr.month, f"MES {curr.month}")
            pal = month_palette[month_index % len(month_palette)]
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            lc = ws.cell(row=row, column=1, value=f"  {mn} {YEAR}")
            lc.font = Font(bold=True, size=13, color=WHITE)
            lc.fill = PatternFill(
                start_color=pal[0], end_color=pal[0], fill_type="solid"
            )
            lc.alignment = Alignment(horizontal="left", vertical="center")
            lc.border = thin
            ws.cell(row=row, column=2).border = thin
            ws.row_dimensions[row].height = 28
            for col in range(3, 3 + len(PEOPLE)):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=pal[1], end_color=pal[1], fill_type="solid"
                )
                ws.cell(row=row, column=col).border = thin
            row += 1

        cal_map.append(row)
        dn = day_n[curr.weekday()]
        pal = month_palette[month_index % len(month_palette)]

        dc = ws.cell(row=row, column=1, value=curr)
        dc.number_format = "DD/MM/YYYY"
        dc.alignment = center
        dc.border = thin
        dc.fill = PatternFill(start_color=pal[2], end_color=pal[2], fill_type="solid")

        dy = ws.cell(row=row, column=2, value=dn)
        dy.alignment = center
        dy.border = thin
        if curr.weekday() == 6:
            dy.font = Font(color="E74C3C", bold=True)
        if curr.weekday() >= 5:
            dc.fill = PatternFill(
                start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid"
            )
            dy.fill = PatternFill(
                start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid"
            )

        for col in range(3, 3 + len(PEOPLE)):
            ws.cell(row=row, column=col).alignment = center
            ws.cell(row=row, column=col).border = thin

        curr += timedelta(days=1)
        row += 1

        # Advance month index
        if month_index + 1 < len(MONTHS) and curr.month > MONTHS[month_index]:
            month_index += 1

    last_cal_row = row - 1
    ws.freeze_panes = "B3"

    # Data validation
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(CODES)}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valor no valido",
        error=f"Usa: {', '.join(CODES)}",
    )
    dv.prompt = f"Selecciona: {' / '.join(CODES)}"
    dv.promptTitle = "Disponibilidad"
    ws.add_data_validation(dv)
    dv.add(f"C2:{get_column_letter(2 + len(PEOPLE))}{last_cal_row}")

    # Conditional formatting for codes
    data_range = f"C2:{get_column_letter(2 + len(PEOPLE))}{last_cal_row}"
    for code, (bg, fg, bold) in CODE_COLORS.items():
        fl = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ft = (
            Font(bold=bold, color=fg, size=10)
            if bold
            else Font(bold=False, color=fg, size=10)
        )
        if code in ("TM", "TT", "FV"):
            ws.conditional_formatting.add(
                data_range, CellIsRule(operator="equal", formula=[f'"{code}"'], fill=fl)
            )
        else:
            ws.conditional_formatting.add(
                data_range, FormulaRule(formula=[f'C2="{code}"'], fill=fl, font=ft)
            )

    # Legend
    lr = last_cal_row + 3
    ws.cell(row=lr, column=1, value="LEYENDA").font = Font(
        bold=True, size=12, color=DARK
    )
    for i, code in enumerate(CODES):
        bg, fg, _ = CODE_COLORS[code]
        txt = f"{code} -> {CODE_MEANINGS.get(code, '')}"
        c = ws.cell(row=lr + 1 + i, column=2, value=txt)
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.font = Font(size=10, bold=True, color=fg)
        c.border = thin
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=lr + 1 + i, start_column=2, end_row=lr + 1 + i, end_column=5
        )
    # Add "vacío = disponible"
    c = ws.cell(row=lr + 1 + len(CODES), column=2, value="(vacio = disponible)")
    c.font = Font(size=10, italic=True, color=DARK)

    # ============================================================
    # SHEET 2: RESUMEN
    # ============================================================
    ws2 = wb.create_sheet("Resumen")
    ws2.sheet_properties.tabColor = "27AE60"
    DSTART = 6

    n_people = len(PEOPLE)
    threshold_cols = min(4, n_people)

    res_widths = [("A", 14), ("B", 7)]
    # Count cols for TM,TT,TN,FV,FN + others? Just show first 5 tracked codes
    tracked = ["TM", "TT", "TN", "FV", "FN"]
    for _ in tracked:
        res_widths.append((get_column_letter(len(res_widths) + 1), 10))
    res_widths.append((get_column_letter(len(res_widths) + 1), 16))  # Ocupados
    for _ in range(threshold_cols):
        res_widths.append((get_column_letter(len(res_widths) + 1), 13))
    res_widths.append((get_column_letter(len(res_widths) + 1), 14))  # Todos FN
    res_widths.append((get_column_letter(len(res_widths) + 1), 14))  # %
    set_col_widths(ws2, res_widths)

    last_res_col = len(res_widths)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_res_col)
    t = ws2.cell(
        row=1,
        column=1,
        value=f"  PANEL DE DISPONIBILIDAD  —  {' & '.join(MONTH_NAMES.get(m, f'MES {m}') for m in MONTHS)} {YEAR}",
    )
    t.font = Font(bold=True, size=18, color=WHITE)
    t.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 46

    # Metric cards
    cards = []
    cards.append(
        (2, "TOTAL DIAS", str(total_days), "34495E", WHITE, "E8F8F5", "1E8449")
    )
    pos = 5
    for th in range(threshold_cols):
        label = ["TODOS LIBRES", "MAX 1 OCUPADO", "MAX 2 OCUPADOS", "MAX 3 OCUPADOS"][
            th
        ]
        col_letter = get_column_letter(10 + th)  # J, K, L, M
        formula = (
            f'=COUNTIF({col_letter}{DSTART}:{col_letter}{DSTART + total_days - 1},"SI")'
        )
        cards.append((pos, label, formula, "1E8449", WHITE, "D5F5E3", "1E8449"))
        pos += 3

    for sc, lbl, val, lbg, lfg, vbg, vfg in cards:
        lc = ws2.cell(row=2, column=sc, value=lbl)
        lc.font = Font(bold=True, size=9, color=lfg)
        lc.fill = PatternFill(start_color=lbg, end_color=lbg, fill_type="solid")
        lc.alignment = center
        lc.border = thin
        vc = ws2.cell(row=2, column=sc + 1, value=val)
        vc.font = Font(bold=True, size=16, color=vfg)
        vc.fill = PatternFill(start_color=vbg, end_color=vbg, fill_type="solid")
        vc.alignment = center
        vc.border = thin

    ws2.row_dimensions[2].height = 48
    ws2.row_dimensions[3].height = 4

    # Resumen headers
    rh = ["Fecha", "Dia", "Libres"] + tracked + ["Ocupados\n(TM+TT+TN)"]
    for th in range(threshold_cols):
        if th == 0:
            rh.append("Todos\nlibres")
        else:
            rh.append(f"Todos -{th}")
    rh.append("Todos\nFN")
    rh.append("% Disponible")

    h2f = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    for col, h_txt in enumerate(rh, 1):
        c = ws2.cell(row=4, column=col, value=h_txt)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = h2f
        c.alignment = center_wrap
        c.border = thin
    ws2.row_dimensions[4].height = 38

    # Description row
    desc_start = 3 + len(tracked) + 2  # after Libres + tracked codes
    ws2.merge_cells(
        start_row=5,
        start_column=desc_start,
        end_row=5,
        end_column=desc_start + threshold_cols - 1,
    )
    ws2.cell(
        row=5, column=desc_start, value="Umbrales: minimo de personas libres"
    ).font = Font(size=9, italic=True, color="7F8C8D")

    # Data rows
    for i, cal_row in enumerate(cal_map):
        rr = DSTART + i
        sd = ws.cell(row=cal_row, column=1).value
        ca = ws2.cell(row=rr, column=1, value=sd)
        ca.number_format = "DD/MM/YYYY"
        ca.alignment = center
        ca.border = thin
        cb = ws2.cell(row=rr, column=2, value=ws.cell(row=cal_row, column=2).value)
        cb.alignment = center
        cb.border = thin
        if sd.weekday() >= 5:
            ca.fill = PatternFill(
                start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid"
            )
            cb.fill = PatternFill(
                start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid"
            )
        elif i % 2:
            ca.fill = PatternFill(
                start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid"
            )
            cb.fill = PatternFill(
                start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid"
            )

        cols_range = f"{get_column_letter(3)}:{get_column_letter(2 + len(tracked))}"
        ws2.cell(
            row=rr, column=3
        ).value = f'=COUNTIF(Calendario!{get_column_letter(3)}{cal_row}:{get_column_letter(2 + len(PEOPLE))}{cal_row},"")'
        for ti, tc in enumerate(tracked):
            cl = get_column_letter(4 + ti)
            ws2.cell(
                row=rr, column=4 + ti
            ).value = f'=COUNTIF(Calendario!{get_column_letter(3)}{cal_row}:{get_column_letter(2 + len(PEOPLE))}{cal_row},"{tc}")'

        occ_col = 4 + len(tracked)
        ws2.cell(row=rr, column=occ_col).value = f"=D{rr}+E{rr}+F{rr}"

        for th in range(threshold_cols):
            col = occ_col + 1 + th
            needed = n_people - th
            ws2.cell(row=rr, column=col).value = f'=IF(C{rr}>={needed},"SI","")'

        fn_col = occ_col + 1 + threshold_cols
        ws2.cell(row=rr, column=fn_col).value = f'=IF(H{rr}={n_people},"SI","")'

        pct_col = fn_col + 1
        ws2.cell(row=rr, column=pct_col).value = f"=C{rr}/{n_people}"
        ws2.cell(row=rr, column=pct_col).number_format = "0%"

        for col in range(3, pct_col + 1):
            ws2.cell(row=rr, column=col).alignment = center
            ws2.cell(row=rr, column=col).border = thin

    lrd = DSTART + len(cal_map) - 1
    tr = lrd + 1
    ws2.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    tc = ws2.cell(row=tr, column=1, value="  TOTALES")
    tc.font = Font(bold=True, size=11, color=WHITE)
    tc.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    tc.border = thin
    ws2.cell(row=tr, column=2).border = thin
    tf = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
    for col in range(3, fn_col + 1):
        cl = get_column_letter(col)
        cell = ws2.cell(row=tr, column=col)
        cell.value = f"=SUM({cl}{DSTART}:{cl}{lrd})"
        cell.font = Font(bold=True, size=11, color="1E8449")
        cell.fill = tf
        cell.alignment = center
        cell.border = thin
    cell = ws2.cell(row=tr, column=pct_col)
    cell.value = f"=AVERAGE({get_column_letter(pct_col)}{DSTART}:{get_column_letter(pct_col)}{lrd})"
    cell.number_format = "0%"
    cell.font = Font(bold=True, size=11, color="1E8449")
    cell.fill = tf
    cell.alignment = center
    cell.border = thin
    ws2.row_dimensions[tr].height = 30

    # CF for threshold cols
    for th in range(threshold_cols):
        col = occ_col + 1 + th
        cl = get_column_letter(col)
        intensity = max(0, 3 - th) * 20
        g = (
            hex(0x1E8449 - intensity * 0x10000)[2:].zfill(6)
            if th == 0
            else hex(0x27AE60 - (th - 1) * 0x10000)[2:].zfill(6)
        )
        bg_colors = ["1E8449", "27AE60", "82E0AA", "ABEBC6"]
        ws2.conditional_formatting.add(
            f"{cl}{DSTART}:{cl}{lrd}",
            CellIsRule(
                operator="equal",
                formula=['"SI"'],
                fill=PatternFill(
                    start_color=bg_colors[th],
                    end_color=bg_colors[th],
                    fill_type="solid",
                ),
                font=Font(bold=th < 2, color=WHITE if th < 2 else DARK, size=11),
            ),
        )

    fn_cl = get_column_letter(fn_col)
    ws2.conditional_formatting.add(
        f"{fn_cl}{DSTART}:{fn_cl}{lrd}",
        CellIsRule(
            operator="equal",
            formula=['"SI"'],
            fill=PatternFill(
                start_color="E06666", end_color="E06666", fill_type="solid"
            ),
            font=Font(bold=True, color=WHITE),
        ),
    )
    pct_cl = get_column_letter(pct_col)
    ws2.conditional_formatting.add(
        f"{pct_cl}{DSTART}:{pct_cl}{lrd}",
        ColorScaleRule(
            start_type="min",
            start_color="E74C3C",
            mid_type="percentile",
            mid_value=50,
            mid_color="F8C471",
            end_type="max",
            end_color="27AE60",
        ),
    )
    ws2.conditional_formatting.add(
        f"A{DSTART}:B{lrd}",
        FormulaRule(
            formula=[f'OR(B{DSTART}="Sab",B{DSTART}="Dom")'],
            fill=PatternFill(
                start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid"
            ),
        ),
    )
    ws2.freeze_panes = f"C{DSTART}"

    # ============================================================
    # SHEETS: una por mes (listas de nombres)
    # ============================================================
    month_sheets = {}
    for mi, m in enumerate(MONTHS):
        mn = MONTH_NAMES.get(m, f"MES {m}")
        pal = month_palette[mi % len(month_palette)]
        ms = wb.create_sheet(mn.capitalize())
        ms.sheet_properties.tabColor = pal[0]
        for cl, w in [("A", 6), ("B", 14), ("C", 7), ("D", 55), ("E", 55)]:
            ms.column_dimensions[cl].width = w

        style_header(
            ms,
            ["Dia", "Fecha", "Dia", "Disponibles (libres)", "Trabajan / Ocupados"],
            row=1,
            fill=PatternFill(start_color=pal[3], end_color=pal[3], fill_type="solid"),
        )
        ms.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ms.cell(
            row=1, column=1, value=f"  {mn} {YEAR}  —  Quien puede cada dia"
        ).font = Font(bold=True, size=13, color=WHITE)
        ms.cell(row=1, column=1).fill = PatternFill(
            start_color=pal[3], end_color=pal[3], fill_type="solid"
        )

        m_start = sum(month_dates.get(MONTHS[j], 0) for j in range(mi))
        m_end = m_start + month_dates.get(m, 0)
        month_entries = cal_map[m_start:m_end]

        for i, cal_row in enumerate(month_entries):
            rr = 2 + i
            sd = ws.cell(row=cal_row, column=1).value
            ms.cell(row=rr, column=1, value=i + 1).alignment = center
            ms.cell(row=rr, column=1).border = thin
            ca = ms.cell(row=rr, column=2, value=sd)
            ca.number_format = "DD/MM/YYYY"
            ca.alignment = center
            ca.border = thin
            cb = ms.cell(row=rr, column=3, value=ws.cell(row=cal_row, column=2).value)
            cb.alignment = center
            cb.border = thin

            av = []
            bv = []
            for idx, p in enumerate(PEOPLE):
                cl = get_column_letter(3 + idx)
                av.append(f'IF(Calendario!{cl}{cal_row}="","{p}","")')
                bv.append(
                    f'IF(Calendario!{cl}{cal_row}<>"","{p}("&Calendario!{cl}{cal_row}&")","")'
                )

            ms.cell(row=rr, column=4).value = f'=TEXTJOIN(", ";TRUE,{", ".join(av)})'
            ms.cell(row=rr, column=4).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ms.cell(row=rr, column=4).border = thin

            ms.cell(row=rr, column=5).value = f'=TEXTJOIN(", ";TRUE,{", ".join(bv)})'
            ms.cell(row=rr, column=5).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ms.cell(row=rr, column=5).border = thin

            if sd.weekday() >= 5:
                for col in range(1, 6):
                    ms.cell(row=rr, column=col).fill = PatternFill(
                        start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid"
                    )
            ms.row_dimensions[rr].height = 22
        ms.freeze_panes = "D2"
        month_sheets[m] = ms

    # ============================================================
    # SHEET: CUADRICULA (grid con emojis)
    # ============================================================
    ws5 = wb.create_sheet("Cuadricula")
    ws5.sheet_properties.tabColor = "8E44AD"
    for ci in range(1, 8):
        ws5.column_dimensions[get_column_letter(ci)].width = 20

    grid_months = []
    for mi, m in enumerate(MONTHS):
        mn = MONTH_NAMES.get(m, f"MES {m}")
        pal = month_palette[mi % len(month_palette)]

        # first day of month column (0=Mon)
        first_date = date(YEAR, m, 1)
        start_col = first_date.weekday()

        m_start = sum(month_dates.get(MONTHS[j], 0) for j in range(mi))
        grid_months.append(
            {
                "name": f"{mn} {YEAR}",
                "start_col": start_col,
                "num_days": month_dates.get(m, 30),
                "cal_start": m_start,
                "color": pal[0],
                "header_row": 1 + mi * 9,
            }
        )

    for idx_gm, gm in enumerate(grid_months):
        hr = gm["header_row"]
        sc = gm["start_col"]

        ws5.merge_cells(start_row=hr, start_column=1, end_row=hr, end_column=7)
        hc = ws5.cell(hr, 1, value=f"  {gm['name']}")
        hc.font = Font(bold=True, size=14, color=WHITE)
        hc.fill = PatternFill(
            start_color=gm["color"], end_color=gm["color"], fill_type="solid"
        )
        hc.alignment = Alignment(horizontal="left", vertical="center")
        hc.border = thin
        for c in range(2, 8):
            ws5.cell(hr, c).border = thin
        ws5.row_dimensions[hr].height = 32

        for ci, dh in enumerate(day_n.values(), 1):
            c = ws5.cell(hr + 1, ci, value=dh)
            c.font = Font(bold=True, size=10, color=WHITE)
            c.fill = PatternFill(
                start_color="34495E", end_color="34495E", fill_type="solid"
            )
            c.alignment = center
            c.border = thin
        ws5.row_dimensions[hr + 1].height = 26

        for week in range(6):
            gr = hr + 2 + week
            ws5.row_dimensions[gr].height = 130
            for dc in range(7):
                dn = week * 7 + dc - sc + 1
                if 1 <= dn <= gm["num_days"]:
                    ci = gm["cal_start"] + dn - 1
                    cr = cal_map[ci]

                    parts = [f'TEXT(Calendario!A{cr},"DD")']
                    for idx, person in enumerate(PEOPLE):
                        cl = get_column_letter(3 + idx)
                        parts.append(
                            f'IF(Calendario!{cl}{cr}="";"{EMOJI["free"]} {person}";"")'
                        )
                        for code in sorted(SHOWN_IN_GRID):
                            parts.append(
                                f'IF(Calendario!{cl}{cr}="{code}";"{EMOJI[code]} {person}({code})";"")'
                            )

                    formula = f"=TEXTJOIN(CHAR(10);TRUE;{','.join(parts)})"
                    cell = ws5.cell(gr, dc + 1, value=formula)
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
                    cell.border = thin
                    cell.font = Font(size=9, color=DARK)

                    sd_ = ws.cell(row=cr, column=1).value
                    if sd_.weekday() >= 5:
                        cell.fill = PatternFill(
                            start_color=WEEKEND_BG,
                            end_color=WEEKEND_BG,
                            fill_type="solid",
                        )
                else:
                    cell = ws5.cell(gr, dc + 1)
                    cell.border = thin
                    cell.fill = PatternFill(
                        start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
                    )

        # CF
        data_top = hr + 2
        data_bot = data_top + 5
        drng = f"A{data_top}:G{data_bot}"
        for emoji, bg in [
            ("🔵", "D6EAF8"),
            ("🟠", "FDEBD0"),
            ("🟣", "E8DAEF"),
            ("🟤", "F5CBA7"),
            ("🟢", "D5F5E3"),
            ("🟡", "FEF9E7"),
        ]:
            ws5.conditional_formatting.add(
                drng,
                FormulaRule(
                    formula=[f'REGEXMATCH(A{data_top},"{emoji}")'],
                    fill=PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
                ),
            )
        for ci in [6, 7]:
            cg = get_column_letter(ci)
            ws5.conditional_formatting.add(
                f"{cg}{data_top}:{cg}{data_bot}",
                FormulaRule(
                    formula=["TRUE"],
                    fill=PatternFill(
                        start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid"
                    ),
                ),
            )

    ws5.freeze_panes = "A2"

    # ============================================================
    # SHEET: INSTRUCCIONES + APPS SCRIPT
    # ============================================================
    ws6 = wb.create_sheet("Instrucciones")
    ws6.sheet_properties.tabColor = "F39C12"
    ws6.column_dimensions["A"].width = 95

    _person_options = "".join(f"<option>{p}</option>" for p in PEOPLE)
    _code_options = "".join(
        f'<option value="{c}">{c} - {CODE_MEANINGS.get(c, c)}</option>' for c in CODES
    )

    gs_code = f"""function onOpen() {{
      SpreadsheetApp.getUi()
        .createMenu('📅 Grupo')
        .addItem('Marcar rango de dias...', 'showDialog')
        .addSeparator()
        .addItem('Limpiar todo', 'clearAll')
        .addToUi();
    }}

    function showDialog() {{
      var html = HtmlService.createHtmlOutput(
        '<style>' +
        'body{{font-family:Arial,sans-serif;padding:20px}}' +
        'label{{display:block;margin-top:12px;font-weight:bold;font-size:13px}}' +
        'select,input{{width:100%;padding:8px;margin-top:4px;border:1px solid #ccc;border-radius:4px;box-sizing:border-box}}' +
        'button{{margin-top:20px;padding:10px 20px;background:#27AE60;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:14px}}' +
        'button:hover{{background:#1E8449}}' +
        '</style>' +
        '<h3>Marcar disponibilidad en lote</h3>' +
        '<label>Persona</label>' +
        '<select id="persona">' +
        '{_person_options}' +
        '</select>' +
        '<label>Codigo</label>' +
        '<select id="codigo">' +
        '{_code_options}' +
        '</select>' +
        '<label>Fecha inicio</label>' +
        '<input type="date" id="inicio">' +
        '<label>Fecha fin</label>' +
        '<input type="date" id="fin">' +
        '<button onclick="marcar()">Aplicar</button>' +
        '<div id="resultado" style="margin-top:15px;padding:10px;border-radius:4px;display:none"></div>' +
        '<script>' +
        'function marcar(){{var p=document.getElementById("persona").value;' +
        'var c=document.getElementById("codigo").value;' +
        'var i=document.getElementById("inicio").value;' +
        'var f=document.getElementById("fin").value;' +
        'if(!i||!f){{var r=document.getElementById("resultado");' +
        'r.style.display="block";r.style.background="#FDEDEC";r.style.color="#E74C3C";' +
        'r.innerHTML="Selecciona fecha de inicio y fin";return}}' +
        'google.script.run.withSuccessHandler(function(msg){{' +
        'var r=document.getElementById("resultado");' +
        'r.style.display="block";r.style.background="#D5F5E3";r.style.color="#1E8449";' +
        'r.innerHTML=msg}}).marcarRango(p,c,i,f)}}' +
        '</script>'
      ).setWidth(380).setHeight(440);
      SpreadsheetApp.getUi().showModalDialog(html, '📅 Marcar rango');
    }}

    function marcarRango(persona, codigo, inicio, fin) {{
      var sheet = SpreadsheetApp.getActiveSpreadsheet().getSheetByName('Calendario');
      if (!sheet) return 'Error: hoja Calendario no encontrada';
      var data = sheet.getDataRange().getValues();
      var headers = data[0];
      var col = -1;
      for (var h = 0; h < headers.length; h++) {{
        if (String(headers[h]).toUpperCase() === persona) {{ col = h + 1; break; }}
      }}
      if (col < 3) return 'Error: persona "' + persona + '" no encontrada';
      var startDate = new Date(inicio);
      var endDate = new Date(fin);
      endDate.setDate(endDate.getDate() + 1);
      var count = 0;
      for (var i = 1; i < data.length; i++) {{
        var cellDate = data[i][0];
        if (cellDate instanceof Date && cellDate >= startDate && cellDate < endDate) {{
          sheet.getRange(i + 1, col).setValue(codigo);
          count++;
        }}
      }}
      return 'Actualizados ' + count + ' dias como ' + codigo + ' para ' + persona;
    }}

    function clearAll() {{
      var ui = SpreadsheetApp.getUi();
      var r = ui.alert('Limpiar todo',
        'Esto borrara TODAS las celdas de disponibilidad. Continuar?',
        ui.ButtonSet.YES_NO);
      if (r === ui.Button.YES) {{
        var sheet = SpreadsheetApp.getActiveSpreadsheet().getSheetByName('Calendario');
        var lastRow = sheet.getLastRow();
        var lastCol = sheet.getLastColumn();
        sheet.getRange('C2:' + sheet.getRange(1, lastCol).getA1Notation().replace(/[0-9]/g,'') + lastRow).clearContent();
        ui.alert('Listo', 'Calendario limpiado', ui.ButtonSet.OK);
      }}
    }}"""

    def L(t, f=None):
        lines.append((t, f))

    lines = []
    L(
        f"CALENDARIO DE DISPONIBILIDAD  —  {' & '.join(MONTH_NAMES.get(m, f'MES {m}') for m in MONTHS)} {YEAR}",
        Font(bold=True, size=18, color=DARK),
    )
    L("")
    L("COMO USARLO:", Font(bold=True, size=13, color="2980B9"))
    L("")
    L(
        "1. En 'Calendario', cada persona rellena su columna con el desplegable:",
        Font(size=11),
    )
    L(
        f"   {' | '.join(f'{c}({CODE_MEANINGS.get(c, c)[:8]})' for c in CODES)}",
        Font(size=11),
    )
    L("   (vacio = disponible)", Font(size=11))
    L("")
    L("2. Las demas hojas se actualizan SOLAS:", Font(size=11))
    L("   Resumen -> metricas del grupo", Font(size=11))
    for m in MONTHS:
        L(
            f"   {MONTH_NAMES.get(m, f'MES {m}')} -> listado de quien puede cada dia",
            Font(size=11),
        )
    L("   Cuadricula -> vista calendario con emojis de colores:", Font(size=11))
    L(
        f"      {EMOJI['free']} = libre    {EMOJI['TM']} = TM    {EMOJI['TT']} = TT    {EMOJI['TN']} = TN    {EMOJI['FV']} = FV",
        Font(size=11),
    )
    L("   (FN/OC/RE/OT no aparecen = 100% no disponibles)", Font(size=11))
    L("")
    L("APPS SCRIPT", Font(bold=True, size=13, color="2980B9"))
    L("")
    L("Instalacion (solo TU, luego funciona para todos):", Font(size=11))
    L("1. Extensiones -> Apps Script", Font(size=11))
    L("2. Borra el codigo y pega el de abajo", Font(size=11))
    L("3. Ctrl+S (guardar) -> permisos (1a vez)", Font(size=11))
    L("4. Vuelve a la hoja, recarga (F5) -> menu '📅 Grupo'", Font(size=11))
    L("")
    L("PERMISOS (pantalla de 'no seguro'):", Font(bold=True, size=11, color="E74C3C"))
    L(
        "   Revisar permisos -> Elegir cuenta -> 'Avanzado' -> 'Ir a [proyecto] (no seguro)' -> 'Permitir'",
        Font(size=11),
    )
    L("")
    L(
        "NOTA: el script SE PIERDE si subes el .xlsx como nuevo archivo.",
        Font(bold=True, size=11),
    )
    L("  Guarda 'codigo_apps_script.gs' por si necesitas re-importar.", Font(size=11))
    L("")
    L("=== CODIGO APPS SCRIPT ===", Font(bold=True, size=12, color=DARK))
    L("")
    for line in gs_code.split("\n"):
        L(line, Font(size=10, name="Consolas", color=DARK))
    L("")
    L("=== FIN ===", Font(bold=True, size=12, color=DARK))

    for i, (txt, ft) in enumerate(lines):
        r = i + 1
        c = ws6.cell(row=r, column=1, value=txt)
        if ft:
            c.font = ft

    # ============================================================
    # GUARDAR
    # ============================================================
    import os

    script_dir = os.path.dirname(os.path.abspath(__file__))

    first = MONTH_NAMES.get(MONTHS[0], str(MONTHS[0])).lower()
    last = MONTH_NAMES.get(MONTHS[-1], str(MONTHS[-1])).lower()
    xlsx_path = os.path.join(
        script_dir,
        f"calendario_{first}-{last}_{group_slug}_{YEAR}.xlsx",
    )
    gs_path = os.path.join(script_dir, f"codigo_apps_script_{group_slug}.gs")

    wb.save(xlsx_path)
    with open(gs_path, "w") as f:
        f.write(gs_code)

    print(f"OK — {os.path.basename(xlsx_path)}")
    print(f"OK — {os.path.basename(gs_path)}")
    print(
        f"\n{total_days} dias | {len(PEOPLE)} personas | {len(CODES)} codigos | Grupo: {group_name}"
    )
    print(f"Meses: {', '.join(MONTH_NAMES.get(m, f'MES {m}') for m in MONTHS)}")
    print(
        f"Hojas: Calendario, Resumen, {', '.join(MONTH_NAMES.get(m, m) for m in MONTHS)}, Cuadricula, Instrucciones"
    )


if __name__ == "__main__":
    generate(YEAR, MONTHS, MONTH_NAMES, PEOPLE, group_name="grupo")
